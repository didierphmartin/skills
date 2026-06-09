"""
Create a new .xlsx from a JSON spec via openpyxl (pure-Python, Pyodide-friendly).

Spec shape is documented in SKILL.md. Supports multiple sheets, formulas
(written as text — Excel/LibreOffice recalculates on first open), header
styling, column widths, frozen panes, cell merges, number formats, and
basic charts (bar/line/pie/area).

Usage:
    python create.py -i spec.json -o output.xlsx
"""
import argparse
import json
import re
import sys
from datetime import datetime, date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.chart import BarChart, LineChart, PieChart, AreaChart, Reference


CHART_CLASSES = {
    "bar": BarChart,
    "line": LineChart,
    "pie": PieChart,
    "area": AreaChart,
}

ALIGN_HORIZ = {"left", "center", "right", "justify"}
ALIGN_VERT = {"top", "center", "bottom"}

ISO_DATE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
ISO_DATETIME = re.compile(r"^\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}(:\d{2})?$")


def coerce_value(v):
    """Convert a JSON-ish value into an openpyxl-friendly cell value."""
    if v is None:
        return None
    if isinstance(v, (int, float, bool)):
        return v
    if isinstance(v, str):
        # Formula passthrough.
        if v.startswith("="):
            return v
        # ISO date / datetime auto-parse.
        if ISO_DATE.match(v):
            try:
                return datetime.strptime(v, "%Y-%m-%d").date()
            except ValueError:
                return v
        if ISO_DATETIME.match(v):
            for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S",
                        "%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M"):
                try:
                    return datetime.strptime(v, fmt)
                except ValueError:
                    continue
            return v
        return v
    if isinstance(v, (datetime, date)):
        return v
    # Lists / dicts → JSON-stringify so Excel shows something readable.
    return json.dumps(v)


def build_font(style):
    s = style or {}
    kw = {}
    if "bold" in s:         kw["bold"] = bool(s["bold"])
    if "italic" in s:       kw["italic"] = bool(s["italic"])
    if "underline" in s:    kw["underline"] = "single" if s["underline"] else None
    if "font_size_pt" in s:
        try: kw["size"] = float(s["font_size_pt"])
        except (TypeError, ValueError): pass
    if "font_name" in s:    kw["name"] = s["font_name"]
    if "color_rgb" in s:    kw["color"] = "FF" + str(s["color_rgb"]).lstrip("#")
    return Font(**kw) if kw else None


def build_fill(style):
    s = style or {}
    if "fill_rgb" not in s:
        return None
    color = "FF" + str(s["fill_rgb"]).lstrip("#")
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def build_alignment(style):
    s = style or {}
    kw = {}
    if "alignment" in s:
        a = str(s["alignment"]).lower()
        if a in ALIGN_HORIZ:
            kw["horizontal"] = a
    if "valign" in s:
        v = str(s["valign"]).lower()
        if v in ALIGN_VERT:
            kw["vertical"] = v
    return Alignment(**kw) if kw else None


def apply_style(cell, style):
    if not style:
        return
    font = build_font(style)
    if font: cell.font = font
    fill = build_fill(style)
    if fill: cell.fill = fill
    align = build_alignment(style)
    if align: cell.alignment = align


def parse_col_range(spec):
    """
    Accept "B", "B:D", or "A1:E10" → list of column letters touched.
    """
    s = (spec or "").strip().upper()
    if not s:
        return []
    # Cell range like A1:E10 → columns A..E
    m = re.match(r"^([A-Z]+)\d+:([A-Z]+)\d+$", s)
    if m:
        a, b = m.group(1), m.group(2)
    else:
        # Letter range like B:D
        m = re.match(r"^([A-Z]+):([A-Z]+)$", s)
        if m:
            a, b = m.group(1), m.group(2)
        else:
            # Single letter
            m = re.match(r"^([A-Z]+)$", s)
            if m:
                return [m.group(1)]
            return []
    start = column_index_from_string(a)
    end = column_index_from_string(b)
    if start > end:
        start, end = end, start
    return [get_column_letter(i) for i in range(start, end + 1)]


def apply_number_formats(ws, fmts):
    if not fmts:
        return
    for spec, fmt in fmts.items():
        cols = parse_col_range(spec)
        if not cols:
            continue
        for col_letter in cols:
            for cell in ws[col_letter]:
                if cell.row == 1:  # leave the header alone
                    continue
                cell.number_format = fmt


def add_chart(ws, chart_spec):
    ctype = (chart_spec.get("type") or "bar").lower()
    cls = CHART_CLASSES.get(ctype)
    if cls is None:
        return
    chart = cls()
    if chart_spec.get("title"):
        chart.title = chart_spec["title"]
    cats_range = chart_spec.get("categories_range")
    vals_range = chart_spec.get("values_range")
    if not vals_range:
        return
    # Reference syntax: Reference(ws, range_string="Sheet1!A1:A10")
    # or Reference(ws, min_col, min_row, max_col, max_row).
    def to_ref(rng):
        # Normalize to Reference via string form.
        return Reference(ws, range_string=f"{ws.title}!{rng}")
    try:
        data_ref = to_ref(vals_range)
        chart.add_data(data_ref, titles_from_data=False)
        if cats_range:
            chart.set_categories(to_ref(cats_range))
    except Exception as e:
        print(f"[create.py] chart skipped — bad range ({e})", file=sys.stderr)
        return
    if "width_in" in chart_spec:
        chart.width = float(chart_spec["width_in"]) * 2.54  # openpyxl uses cm
    if "height_in" in chart_spec:
        chart.height = float(chart_spec["height_in"]) * 2.54
    anchor = chart_spec.get("anchor", "G2")
    ws.add_chart(chart, anchor)


def render_sheet(wb, idx, sheet_spec):
    name = sheet_spec.get("name") or f"Sheet{idx + 1}"
    if idx == 0:
        ws = wb.active
        ws.title = name
    else:
        ws = wb.create_sheet(title=name)

    header = sheet_spec.get("header") or []
    rows = sheet_spec.get("rows") or []

    # Write header row.
    if header:
        for c, val in enumerate(header, start=1):
            cell = ws.cell(row=1, column=c, value=coerce_value(val))
            apply_style(cell, sheet_spec.get("header_style"))

    # Write data rows.
    row_offset = 2 if header else 1
    for r_idx, row in enumerate(rows):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=row_offset + r_idx, column=c_idx, value=coerce_value(val))

    # Column widths.
    for c_idx, width in enumerate(sheet_spec.get("column_widths") or [], start=1):
        try:
            ws.column_dimensions[get_column_letter(c_idx)].width = float(width)
        except (TypeError, ValueError):
            continue

    # Freeze panes.
    if sheet_spec.get("freeze"):
        ws.freeze_panes = str(sheet_spec["freeze"])

    # Merges.
    for rng in sheet_spec.get("merges") or []:
        try:
            ws.merge_cells(str(rng))
        except Exception as e:
            print(f"[create.py] merge skipped ({rng}): {e}", file=sys.stderr)

    # Number formats.
    apply_number_formats(ws, sheet_spec.get("number_formats"))

    # Charts.
    for chart_spec in sheet_spec.get("charts") or []:
        if isinstance(chart_spec, dict):
            add_chart(ws, chart_spec)


def apply_metadata(wb, meta):
    cp = wb.properties
    if "title" in meta:    cp.title = str(meta["title"])
    if "author" in meta:   cp.creator = str(meta["author"])
    if "subject" in meta:  cp.subject = str(meta["subject"])
    if "keywords" in meta: cp.keywords = str(meta["keywords"])


def main():
    parser = argparse.ArgumentParser(description="Create .xlsx from JSON spec")
    parser.add_argument("-i", "--input", required=True)
    parser.add_argument("-o", "--output", required=True)
    args = parser.parse_args()

    spec_path = Path(args.input)
    out_path = Path(args.output)
    if not spec_path.exists():
        print(f"Spec file not found: {spec_path}", file=sys.stderr)
        sys.exit(2)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    spec = json.loads(spec_path.read_text(encoding="utf-8"))

    wb = Workbook()
    if "metadata" in spec:
        apply_metadata(wb, spec["metadata"])

    sheets = spec.get("sheets") or []
    if not sheets:
        sheets = [{"name": "Sheet1"}]
    for idx, sheet_spec in enumerate(sheets):
        if isinstance(sheet_spec, dict):
            render_sheet(wb, idx, sheet_spec)

    wb.save(out_path)
    print(f"Wrote {out_path.stat().st_size} bytes to {out_path}")
    print(f"Created {len(wb.sheetnames)} sheet(s): {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
