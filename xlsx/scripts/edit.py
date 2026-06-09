"""
Edit an existing .xlsx — find/replace plus targeted cell ops (set_cells,
append_rows, delete_rows). Operations apply in declared order.

Usage:
    python edit.py -i input.xlsx -o output.xlsx --ops ops.json
"""
import argparse
import json
import re
import sys
from datetime import datetime, date
from pathlib import Path

from openpyxl import load_workbook


ISO_DATE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
ISO_DATETIME = re.compile(r"^\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}(:\d{2})?$")


def coerce_value(v):
    if v is None:
        return None
    if isinstance(v, (int, float, bool)):
        return v
    if isinstance(v, str):
        if v.startswith("="):
            return v
        if ISO_DATE.match(v):
            try: return datetime.strptime(v, "%Y-%m-%d").date()
            except ValueError: return v
        if ISO_DATETIME.match(v):
            for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S",
                        "%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M"):
                try: return datetime.strptime(v, fmt)
                except ValueError: continue
            return v
        return v
    if isinstance(v, (datetime, date)):
        return v
    return json.dumps(v)


def find_replace(wb, find, replace, case_sensitive):
    """Replace within string cells across all sheets. Skips formulas."""
    if not find:
        return 0
    count = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, str):
                    continue
                if v.startswith("="):
                    continue   # don't rewrite formulas via find/replace
                if case_sensitive:
                    if find in v:
                        cell.value = v.replace(find, replace)
                        count += 1
                else:
                    new = re.sub(re.escape(find), replace, v, flags=re.IGNORECASE)
                    if new != v:
                        cell.value = new
                        count += 1
    return count


def get_sheet(wb, name):
    if not name:
        return wb.active
    if name in wb.sheetnames:
        return wb[name]
    print(f"[edit.py] sheet not found: {name}", file=sys.stderr)
    return None


def set_cells(wb, ops):
    n = 0
    for op in ops:
        if not isinstance(op, dict):
            continue
        ws = get_sheet(wb, op.get("sheet"))
        if ws is None:
            continue
        cell_ref = op.get("cell")
        if not cell_ref:
            continue
        try:
            ws[cell_ref] = coerce_value(op.get("value"))
            n += 1
        except Exception as e:
            print(f"[edit.py] set_cell skipped ({cell_ref}): {e}", file=sys.stderr)
    return n


def append_rows(wb, ops):
    n = 0
    for op in ops:
        if not isinstance(op, dict):
            continue
        ws = get_sheet(wb, op.get("sheet"))
        if ws is None:
            continue
        row = op.get("row") or []
        if not isinstance(row, list):
            continue
        ws.append([coerce_value(v) for v in row])
        n += 1
    return n


def delete_rows(wb, ops):
    n = 0
    for op in ops:
        if not isinstance(op, dict):
            continue
        ws = get_sheet(wb, op.get("sheet"))
        if ws is None:
            continue
        try:
            start = int(op.get("from", 0))
            count = int(op.get("count", 1))
            if start <= 0 or count <= 0:
                continue
            ws.delete_rows(start, count)
            n += count
        except Exception as e:
            print(f"[edit.py] delete_rows skipped: {e}", file=sys.stderr)
    return n


def main():
    parser = argparse.ArgumentParser(description="Edit an existing .xlsx")
    parser.add_argument("-i", "--input", required=True)
    parser.add_argument("-o", "--output", required=True)
    parser.add_argument("--ops", required=True)
    args = parser.parse_args()

    in_path = Path(args.input)
    out_path = Path(args.output)
    ops_path = Path(args.ops)
    for p, label in [(in_path, "input"), (ops_path, "ops")]:
        if not p.exists():
            print(f"{label.capitalize()} file not found: {p}", file=sys.stderr)
            sys.exit(2)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    ops = json.loads(ops_path.read_text(encoding="utf-8"))
    keep_vba = in_path.suffix.lower() == ".xlsm"
    wb = load_workbook(in_path, keep_vba=keep_vba)

    fr_count = 0
    for op in ops.get("find_replace", []) or []:
        if not isinstance(op, dict):
            continue
        fr_count += find_replace(
            wb,
            op.get("find"),
            op.get("replace", ""),
            bool(op.get("case_sensitive", False)),
        )

    set_count = set_cells(wb, ops.get("set_cells") or [])
    append_count = append_rows(wb, ops.get("append_rows") or [])
    del_count = delete_rows(wb, ops.get("delete_rows") or [])

    wb.save(out_path)
    summary = (
        f"Wrote {out_path.stat().st_size} bytes to {out_path}. "
        f"find/replace={fr_count}, set_cells={set_count}, "
        f"appended={append_count}, deleted_rows={del_count}."
    )
    print(summary)


if __name__ == "__main__":
    main()
